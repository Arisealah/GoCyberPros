# #!/usr/bin/env python3
# """
# XLSX Metadata Extraction Tool (Enhanced)

# Description: Comprehensive and robust metadata extraction from Excel XLSX files,
# with advanced validation, detailed error handling, and standardized, enriched output.

# Features:
# - Extracts core properties, custom properties, sheet information.
# - Validates XLSX structure and integrity rigorously (ZIP+XML).
# - Handles modern Excel formats.
# - Outputs standardized, structured metadata including warnings and errors.

# Dependencies:
# - openpyxl (pip install openpyxl)
# - python-magic (pip install python-magic)

# Example Usage (CLI):
#     python3 xlsx_extractor.py example.xlsx --log-level DEBUG -o metadata.json

# Author: Gemini (based on provided foundation)
# Version: 1.1.0
# Last Updated: 2025-06-07
# """

# import os
# import zipfile
# import magic
# import mimetypes # For more robust MIME type handling
# from datetime import datetime, timezone
# from typing import Dict, Any, Optional, List
# from openpyxl import load_workbook
# import logging
# import sys
# import argparse
# import time
# import json

# # --- Configuration ---
# DEFAULT_LOG_FILE = "xlsx_extractor.log"
# DEFAULT_LOG_LEVEL = "INFO" # DEBUG, INFO, WARNING, ERROR, CRITICAL

# # Expected XLSX MIME types
# EXPECTED_XLSX_MIMES = [
#     'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
#     'application/zip' # XLSX files are essentially ZIP files
# ]

# # --- Logging Setup ---
# def setup_logging(log_level: str = DEFAULT_LOG_LEVEL):
#     """Configures the logging for the extractor."""
#     log_formatter = logging.Formatter('%(asctime)s [%(levelname)s] [%(name)s] %(message)s')
#     root_logger = logging.getLogger()
#     root_logger.setLevel(getattr(logging, log_level.upper()))

#     # Clear existing handlers to prevent duplicate output when called multiple times
#     if root_logger.handlers:
#         for handler in root_logger.handlers:
#             root_logger.removeHandler(handler)

#     # File Handler
#     file_handler = logging.FileHandler(DEFAULT_LOG_FILE, encoding='utf-8')
#     file_handler.setFormatter(log_formatter)
#     root_logger.addHandler(file_handler)

#     # Console Handler
#     console_handler = logging.StreamHandler(sys.stdout)
#     console_handler.setFormatter(log_formatter)
#     root_logger.addHandler(console_handler)

# # Initial logging setup
# setup_logging()
# logger = logging.getLogger(__name__)

# # --- Custom Exception Classes ---
# class XLSXMetadataError(Exception):
#     """Base class for XLSX metadata extraction exceptions."""
#     pass

# class FileAccessError(XLSXMetadataError):
#     """Raised for issues accessing the file (e.g., not found, permissions)."""
#     pass

# class InvalidFileFormatError(XLSXMetadataError):
#     """Raised when the file is not a valid XLSX according to checks."""
#     pass

# class ExtractorProcessingError(XLSXMetadataError):
#     """Raised for unexpected errors during the XLSX extraction process."""
#     pass

# # --- Helper Functions ---
# def _datetime_to_iso_utc(dt_obj: Optional[datetime]) -> Optional[str]:
#     """Converts a datetime object to ISO 8601 format with UTC timezone, if not None."""
#     if dt_obj:
#         if dt_obj.tzinfo is None: # Assume local time if no timezone info
#             return dt_obj.astimezone(timezone.utc).isoformat()
#         return dt_obj.astimezone(timezone.utc).isoformat()
#     return None

# # --- Main Validation Function ---
# def validate_xlsx_file(file_path: str) -> bool:
#     """
#     Advanced validation for XLSX file structure and integrity.

#     Args:
#         file_path: Path to the XLSX file.

#     Returns:
#         True if the file passes all validations.

#     Raises:
#         FileAccessError: If file doesn't exist or access is restricted.
#         InvalidFileFormatError: If file is not deemed a valid XLSX.
#     """
#     logger.info(f"Validating file: {file_path}")

#     if not os.path.exists(file_path):
#         raise FileAccessError(f"File not found: {file_path}")

#     if not os.path.isfile(file_path):
#         raise FileAccessError(f"Path is not a regular file: {file_path}")

#     if not os.access(file_path, os.R_OK):
#         raise FileAccessError(f"Access denied (read permission): {file_path}")

#     file_size = os.path.getsize(file_path)
#     if file_size == 0:
#         logger.warning(f"File {file_path} is empty, which is an invalid XLSX.")
#         raise InvalidFileFormatError(f"File is empty: {file_path}")

#     # 1. MIME Type Check (using both magic and mimetypes)
#     mime_type_magic = magic.from_file(file_path, mime=True)
#     mime_type_ext = mimetypes.guess_type(file_path)[0]

#     is_mime_ok = mime_type_magic in EXPECTED_XLSX_MIMES
#     if not is_mime_ok:
#         if mime_type_ext and mime_type_ext in EXPECTED_XLSX_MIMES:
#             logger.warning(f"Magic detected '{mime_type_magic}', but extension suggests XLSX ('{mime_type_ext}'). Proceeding.")
#         else:
#             raise InvalidFileFormatError(f"Not a recognized XLSX file type by magic ({mime_type_magic}).")
    
#     # 2. ZIP structure and essential XLSX file check
#     try:
#         with zipfile.ZipFile(file_path) as z:
#             # Check for core XLSX components
#             required_files = ['xl/workbook.xml', 'xl/worksheets/sheet1.xml', '_rels/.rels']
#             for req_file in required_files:
#                 if req_file not in z.namelist():
#                     raise InvalidFileFormatError(f"Missing required XLSX internal file: {req_file}")
            
#             # Check for common corruption indicators (e.g., empty or invalid central directory)
#             if not z.namelist():
#                 raise InvalidFileFormatError("XLSX zip archive is empty or corrupted.")

#     except zipfile.BadZipFile as e:
#         raise InvalidFileFormatError(f"Invalid ZIP structure for XLSX: {str(e)}") from e
#     except Exception as e:
#         logger.error(f"Unexpected error during ZIP validation for {file_path}: {e}", exc_info=True)
#         raise InvalidFileFormatError(f"Unexpected validation error: {str(e)}") from e

#     logger.info(f"File '{file_path}' passed all structural validations.")
#     return True

# # --- Main Extraction Function ---
# def extract_xlsx_metadata(file_path: str) -> Dict[str, Any]:
#     """
#     Extract comprehensive metadata from XLSX files.

#     Args:
#         file_path: Path to the XLSX file.

#     Returns:
#         A dictionary containing extracted metadata, including:
#         - file_info: Basic file system attributes.
#         - core_properties: Standard XLSX document properties.
#         - workbook_info: Sheet and workbook structure data.
#         - processing: Status, warnings, errors, and time taken.

#     Raises:
#         XLSXMetadataError: For any critical failure during validation or processing.
#     """
#     result: Dict[str, Any] = {
#         "file_info": {},
#         "core_properties": {},
#         "workbook_info": {},
#         "processing": {
#             "success": False,
#             "warnings": [],
#             "errors": [],
#             "time_taken_seconds": None,
#             "extractor_version": "1.1.0"
#         }
#     }

#     start_time = time.time()
#     logger.info(f"Starting XLSX metadata extraction for: {file_path}")

#     wb = None # Initialize to None for finally block

#     try:
#         # --- Stage 1: Validation ---
#         validate_xlsx_file(file_path)
#         result["file_info"]["valid"] = True

#         # --- Stage 2: File System Info ---
#         file_stat = os.stat(file_path)
#         result["file_info"] = {
#             "path": os.path.abspath(file_path),
#             "filename": os.path.basename(file_path),
#             "size_bytes": file_stat.st_size,
#             "created_utc": datetime.fromtimestamp(file_stat.st_ctime, tz=timezone.utc).isoformat(),
#             "modified_utc": datetime.fromtimestamp(file_stat.st_mtime, tz=timezone.utc).isoformat(),
#             "accessed_utc": datetime.fromtimestamp(file_stat.st_atime, tz=timezone.utc).isoformat(),
#             "format": "XLSX",
#             "detected_mime_type": mimetypes.guess_type(file_path)[0] or magic.from_file(file_path, mime=True),
#             "valid": True
#         }

#         # --- Stage 3: XLSX-specific Metadata Extraction (using openpyxl) ---
#         try:
#             # read_only=True for faster loading and less memory usage if not modifying
#             # data_only=True to get displayed cell values, not formulas
#             wb = load_workbook(file_path, read_only=True, data_only=True)
#             props = wb.properties
            
#             # Core Properties
#             result["core_properties"] = {
#                 "title": props.title,
#                 "creator": props.creator,
#                 "description": props.description,
#                 "subject": props.subject,
#                 "keywords": props.keywords,
#                 "last_modified_by": props.lastModifiedBy,
#                 "created_utc": _datetime_to_iso_utc(props.created),
#                 "modified_utc": _datetime_to_iso_utc(props.modified),
#                 "category": props.category,
#                 "revision": props.revision,
#                 "version": props.version # Added version property
#             }
            
#             # Workbook Info
#             result["workbook_info"] = {
#                 "total_sheets": len(wb.sheetnames),
#                 "sheet_names": wb.sheetnames,
#                 "active_sheet_name": wb.active.title if wb.active else None,
#                 "defined_names": list(wb.defined_names),
#                 "has_macros": False # Default to False, check below
#             }

#             # Check for macros (.xlsm)
#             # This is a robust check using zipfile to look for vbaProject.bin
#             with zipfile.ZipFile(file_path, 'r') as zf:
#                 if 'xl/vbaProject.bin' in zf.namelist():
#                     result["workbook_info"]["has_macros"] = True
#                     result["processing"]["warnings"].append("Workbook contains VBA macros.")

#             # Basic sheet statistics (e.g., number of cells, rows, columns in active sheet)
#             if wb.active:
#                 active_sheet = wb.active
#                 result["workbook_info"]["active_sheet_dimensions"] = {
#                     "max_row": active_sheet.max_row,
#                     "max_column": active_sheet.max_column,
#                     "min_row": active_sheet.min_row,
#                     "min_column": active_sheet.min_column,
#                     "total_cells_approx": active_sheet.max_row * active_sheet.max_column
#                 }
#                 # Check for protected sheets
#                 if active_sheet.protection.sheet:
#                      result["workbook_info"]["active_sheet_protected"] = True
#                      result["processing"]["warnings"].append(f"Active sheet '{active_sheet.title}' is protected.")
#                 else:
#                      result["workbook_info"]["active_sheet_protected"] = False
#             else:
#                 result["processing"]["warnings"].append("No active sheet found in workbook.")
                
#             result["processing"]["success"] = True
        
#         except Exception as e:
#             result["processing"]["errors"].append(f"XLSX content read error: {str(e)}. File might be corrupted or malformed internally.")
#             logger.error(f"OpenPyXL processing error: {e}")
#             result["processing"]["success"] = False # Mark as failed
#             return result # Exit early

#     except FileAccessError as e:
#         result["processing"]["errors"].append(str(e))
#         logger.error(f"File Access Error during XLSX extraction: {e}")
#     except InvalidFileFormatError as e:
#         result["processing"]["errors"].append(str(e))
#         logger.error(f"Invalid XLSX Format Error: {e}")
#     except Exception as e:
#         # Catch any other unexpected errors that might occur outside specific blocks
#         result["processing"]["errors"].append(f"An unhandled critical error occurred: {str(e)}")
#         logger.exception(f"An unhandled critical error occurred during extraction for {file_path}")
#     finally:
#         result["processing"]["time_taken_seconds"] = time.time() - start_time
#         if wb:
#             try:
#                 wb.close() # Ensure workbook is closed
#             except Exception as e:
#                 logger.warning(f"Error closing workbook: {e}")
#         if not result["processing"]["success"] and not result["processing"]["errors"]:
#              # If success is false but no errors recorded, it's an unhandled case
#             result["processing"]["errors"].append("Extraction failed without specific error message (check logs).")
#         logger.info(f"Extraction finished for {file_path}. Success: {result['processing']['success']}")

#     return result

# # --- CLI Entry Point ---
# def run_cli():
#     """Command-line interface for the XLSX metadata extractor."""
#     parser = argparse.ArgumentParser(
#         description="Extract comprehensive metadata from an XLSX file. "
#                     "Outputs JSON metadata to stdout or a specified file."
#     )
#     parser.add_argument(
#         "file",
#         nargs="?", # Optional argument for file path
#         help="Path to the XLSX file to extract metadata from."
#     )
#     parser.add_argument(
#         "--log-level",
#         default=DEFAULT_LOG_LEVEL,
#         choices=["DEBUG", "INFO", "WARNING", "ERROR", "CRITICAL"],
#         help=f"Set the logging level (default: {DEFAULT_LOG_LEVEL})."
#     )
#     parser.add_argument(
#         "--output",
#         "-o",
#         help="Write metadata output to this file (JSON). If not set, prints to stdout."
#     )
#     parser.add_argument(
#         "--no-input-prompt",
#         action="store_true",
#         help="Do not prompt for file path if not provided as argument. Exit instead."
#     )

#     args = parser.parse_args()

#     # Re-setup logging based on CLI argument
#     setup_logging(args.log_level)
    
#     file_path = args.file
#     if not file_path:
#         if args.no_input_prompt:
#             logger.error("No file path provided and --no-input-prompt is set. Exiting.")
#             sys.exit(1)
#         else:
#             file_path = input("Enter the path to the XLSX file: ").strip()
#             if not file_path:
#                 logger.error("No file path provided. Exiting.")
#                 sys.exit(1)

#     try:
#         logger.info(f"Attempting to extract metadata for: {file_path}")
#         metadata = extract_xlsx_metadata(file_path)
#         output_json = json.dumps(metadata, indent=2, ensure_ascii=False) # ensure_ascii for proper non-ASCII display

#         if args.output:
#             with open(args.output, "w", encoding="utf-8") as f: # Ensure UTF-8 output
#                 f.write(output_json)
#             logger.info(f"Metadata successfully written to {args.output}")
#             print(f"Metadata written to {args.output}")
#         else:
#             print(output_json)
        
#         # Indicate non-zero exit code if extraction was not fully successful
#         if not metadata["processing"]["success"] or metadata["processing"]["errors"]:
#             logger.warning(f"Extraction completed with warnings/errors for {file_path}. Exit code 2.")
#             sys.exit(2)
#         else:
#             logger.info(f"Extraction successful for: {file_path}. Time taken: {metadata['processing']['time_taken_seconds']:.2f}s")

#     except Exception as e: # Catch any remaining unhandled exceptions from CLI setup or argparse
#         logger.exception(f"A critical error occurred in the CLI for {file_path}: {e}")
#         print(f"A critical error occurred: {e}. Check {DEFAULT_LOG_FILE} for details.")
#         sys.exit(1)

# def main():
#     run_cli()

# if __name__ == "__main__":
#     main()











## version 2 




#!/usr/bin/env python3
"""
XLSX Metadata Extraction Tool

Description: Comprehensive metadata extraction from XLSX files with validation,
error handling, and standardized output.

Features:
- Extracts core properties, sheet information
- Validates XLSX structure (ZIP+XML)
- Handles modern Excel formats
- Command-line interface with user prompts
- Error handling with JSON output
- Option to save output to file
"""

import os
import zipfile
import magic
import argparse
import json
from datetime import datetime, timezone
from typing import Dict, Any, Optional
from openpyxl import load_workbook

class XLSXMetadataError(Exception):
    """Base class for XLSX metadata exceptions"""
    pass

class InvalidXLSXError(XLSXMetadataError):
    """Raised when file is not a valid XLSX"""
    pass

class XLSXProcessingError(XLSXMetadataError):
    """Raised during XLSX processing failures"""
    pass

def validate_xlsx_file(file_path: str) -> bool:
    """
    Validate XLSX file structure and integrity.
    
    Args:
        file_path: Path to the XLSX file
        
    Raises:
        FileNotFoundError: If file doesn't exist
        PermissionError: If file access is restricted
        InvalidXLSXError: If file is not a valid XLSX
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")
    
    if not os.access(file_path, os.R_OK):
        raise PermissionError(f"Access denied: {file_path}")
    
    mime = magic.from_file(file_path, mime=True)
    if mime not in ('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'application/zip'):
        raise InvalidXLSXError(f"Not an XLSX file (detected: {mime})")
    
    # Check ZIP structure
    try:
        with zipfile.ZipFile(file_path) as z:
            required_files = ['xl/workbook.xml', 'xl/worksheets/sheet1.xml']
            for req_file in required_files:
                if req_file not in z.namelist():
                    raise InvalidXLSXError(f"Missing required XLSX file: {req_file}")
    except zipfile.BadZipFile as e:
        raise InvalidXLSXError(f"Invalid ZIP structure: {str(e)}")
    
    return True

def _datetime_to_iso(dt_obj: Optional[datetime]) -> Optional[str]:
    """Convert datetime to ISO format"""
    if dt_obj:
        return dt_obj.astimezone(timezone.utc).isoformat()
    return None

def extract_xlsx_metadata(file_path: str) -> Dict[str, Any]:
    """
    Extract comprehensive metadata from XLSX files.
    
    Returns:
        Dictionary containing:
        - file_info: Basic file attributes
        - core_properties: Document properties
        - workbook_info: Sheet and workbook data
        - processing: Status information
        
    Raises:
        XLSXMetadataError: For any processing failures
    """
    result = {
        "file_info": {},
        "core_properties": {},
        "workbook_info": {},
        "processing": {
            "success": False,
            "warnings": [],
            "time_taken": None
        }
    }
    
    start_time = datetime.now()
    workbook = None
    
    try:
        validate_xlsx_file(file_path)
        
        file_stat = os.stat(file_path)
        result["file_info"] = {
            "path": os.path.abspath(file_path),
            "size_bytes": file_stat.st_size,
            "created": datetime.fromtimestamp(file_stat.st_ctime).isoformat(),
            "modified": datetime.fromtimestamp(file_stat.st_mtime).isoformat(),
            "format": "XLSX",
            "valid": True
        }
        
        # Open workbook in read-only mode
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        props = workbook.properties
        
        # Core properties
        result["core_properties"] = {
            "title": props.title,
            "creator": props.creator,
            "description": props.description,
            "subject": props.subject,
            "keywords": props.keywords,
            "last_modified_by": props.lastModifiedBy,
            "created": _datetime_to_iso(props.created),
            "modified": _datetime_to_iso(props.modified),
            "category": props.category,
            "revision": props.revision
        }
        
        # Workbook info
        result["workbook_info"] = {
            "sheets": len(workbook.sheetnames),
            "sheet_names": workbook.sheetnames,
            "active_sheet": workbook.active.title if workbook.active else None,
            "has_macros": any(name.endswith('.bin') for name in zipfile.ZipFile(file_path).namelist())
        }
        
        # Add active sheet dimensions if available
        if workbook.active:
            sheet = workbook.active
            result["workbook_info"]["active_sheet_dimensions"] = {
                "rows": sheet.max_row,
                "columns": sheet.max_column
            }
        
        result["processing"]["success"] = True
        
    except FileNotFoundError as e:
        raise XLSXMetadataError(f"File error: {str(e)}") from e
    except PermissionError as e:
        raise XLSXMetadataError(f"Access error: {str(e)}") from e
    except zipfile.BadZipFile as e:
        raise InvalidXLSXError(f"Invalid XLSX/ZIP: {str(e)}") from e
    except Exception as e:
        raise XLSXProcessingError(f"Processing failed: {str(e)}") from e
    finally:
        if workbook:
            workbook.close()
        result["processing"]["time_taken"] = (datetime.now() - start_time).total_seconds()
    
    return result

def save_output_to_file(output: Dict, output_path: Optional[str] = None) -> None:
    """Save the output to a file if requested"""
    if output_path:
        try:
            with open(output_path, 'w') as f:
                json.dump(output, f, indent=2)
            print(f"\nMetadata saved to: {output_path}")
        except Exception as e:
            print(f"\nWarning: Could not save output to file: {str(e)}")

def main():
    """Command-line interface for the XLSX metadata extractor"""
    parser = argparse.ArgumentParser(description='XLSX Metadata Extraction Tool')
    parser.add_argument('file_path', nargs='?', help='Path to the XLSX file')
    parser.add_argument('--output', '-o', help='Path to save the output JSON')
    args = parser.parse_args()
    
    file_path = args.file_path
    if not file_path:
        file_path = input("Enter the path to the XLSX file: ")
    
    try:
        metadata = extract_xlsx_metadata(file_path)
        print("\nXLSX Metadata:")
        print(json.dumps(metadata, indent=2))
        
        if args.output:
            save_output_to_file(metadata, args.output)
        
    except XLSXMetadataError as e:
        error_output = {
            "error": str(e),
            "success": False,
            "file": file_path
        }
        print(json.dumps(error_output, indent=2))
        exit(1)

if __name__ == "__main__":
    main()